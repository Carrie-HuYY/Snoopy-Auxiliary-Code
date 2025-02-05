import pandas as pd
import time



def convert_xlsx_to_csv(xlsx_file, csv_file):
    """
    将大型 .xlsx 文件转换为 .csv 文件
    """
    # 读取 .xlsx 文件
    df = pd.read_excel(xlsx_file)
    # 保存为 .csv 文件
    df.to_csv(csv_file, index=False)


def process_large_csv(file_path, chunk_size, search_term):
    """
    分块读取大型Excel文件并进行检索
    :param file_path: 文件路径
    :param chunk_size: 每次读取的行数
    :param search_term: 搜索关键词
    """
    # 使用pandas的read_excel分块读取
    chunks = pd.read_csv(file_path, chunksize=chunk_size)

    # 遍历每个分块
    for chunk in chunks:
        if search_term:
            # 检索包含关键词的行
            filtered_chunk = chunk[chunk.apply(lambda row: row.astype(str).str.contains(search_term).any(), axis=1)]
            if not filtered_chunk.empty:
                print("找到匹配的行：")
                print(filtered_chunk)
        else:
            print(chunk)  # 打印当前分块内容

# 使用示例

import pandas as pd
import os


def process_xlsx_files(folder_path, output_folder):
    """
    处理指定文件夹中的所有.xlsx文件，只保留PMID、Title和Abstract三列，并另存为.csv文件。
    :param folder_path: 包含.xlsx文件的文件夹路径
    :param output_folder: 输出的文件夹路径
    """
    # 确保输出文件夹存在
    os.makedirs(output_folder, exist_ok=True)

    # 遍历文件夹中的所有文件
    for file_name in os.listdir(folder_path):
        if file_name.endswith('.xlsx'):
            file_path = os.path.join(folder_path, file_name)
            try:
                # 读取.xlsx文件
                df = pd.read_excel(file_path)
                # 保留特定列
                if 'PMID' in df.columns and 'Title' in df.columns and 'Abstract' in df.columns:
                    df = df[['PMID', 'Title', 'Abstract']]
                    # 保存为.csv文件
                    output_file = os.path.join(output_folder, file_name.replace('.xlsx', '.xlsx'))
                    df.to_excel(output_file, index=False)
                    print(f"已处理并保存文件：{output_file}")
                else:
                    print(f"文件 {file_name} 中缺少必要的列，跳过处理。")
            except Exception as e:
                print(f"处理文件 {file_name} 时出错：{e}")


from openpyxl import load_workbook

def process_large_xlsx(file_path, search_term=None):
    """
    逐行读取大型 .xlsx 文件并进行检索
    """
    # 加载工作簿
    workbook = load_workbook(file_path, read_only=True)
    sheet = workbook.active  # 获取活动工作表

    # 遍历每一行
    for row in sheet.iter_rows(values_only=True):
        if search_term:
            if any(search_term in str(cell) for cell in row):
                print("找到匹配的行：", row)
        else:
            print(row)  # 打印当前行内容

# 使用示例
start_time = time.time()
folder_path = r"D:\Snoopy\data\pubmed"

file_paths = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.endswith('.xlsx')]

for file_path in file_paths:
    process_large_xlsx(file_path, search_term="HCC")

finish_time = time.time()
print(finish_time - start_time)



